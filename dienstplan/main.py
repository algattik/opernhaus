#!/usr/bin/env python3

import os
import sys
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from collections import defaultdict
from ics import Calendar, Event
import pytz
import uuid


# Global constants
UUID_NAMESPACE = uuid.UUID('B87533E8-8A42-43BD-B60C-7E97EA456009')
DATE_FORMAT = "%d.%m.%Y"
COLOR1 = PatternFill(start_color="00DDFFDD", fill_type="solid")
COLOR2 = PatternFill(start_color="00FFDDDD", fill_type="solid")
TIMEZONE = pytz.timezone('Europe/Zurich')

# Global styles
THIN_BORDER = Border(top=Side(style='thin'), bottom=Side(style='thin'))
THICK_BORDER = Border(top=Side(style='thick'), bottom=Side(style='thin'))
BOLD_FONT = Font(bold=True)
TOP_ALIGNMENT = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Utility function to iterate over days
def iterate_days(start_date, end_date):
    current_date = start_date
    while current_date <= end_date:
        yield current_date
        current_date += timedelta(days=1)

# Function to add a row to a worksheet
def add_row(worksheet, row_data, fill_color):
    worksheet.append(row_data)
    for coln in range(0, len(row_data)):
        cell = worksheet.cell(row=worksheet.max_row, column=coln+1)
        cell.alignment = TOP_ALIGNMENT
        if coln == 0:
            cell.fill = fill_color

def merge_strings(row_data):
    return '\n'.join(row_data)

def contains_substring(input_string, substrings):
    for substring in substrings:
        if substring in input_string:
            return True
    return False

def parse_hour(shift_hours):
    if shift_hours == 'None':
        return '00.00'
    if '.' in shift_hours:
        return shift_hours
    return f'{shift_hours}.00'

def dienst_duration(show_info, points):
    if show_info == "für den DPE":
        return None
    if show_info == "vom Haus":
        return None
    if show_info == "U11":
        return timedelta(days=11)
    if show_info == "SUVA":
        return None
    if show_info == '' and points is None:
        return None
    if points == 1:
        return timedelta(hours=3, minutes=15)
    if points == 1.5:
        return timedelta(hours=4)
    if points == 2:
        return timedelta(hours=5)
    if points == 2.5:
        return timedelta(hours=6)
    if points == 3:
        return timedelta(hours=6, minutes=30)
    raise ValueError(f"Invalid points value: {points}")

# Command-line argument handling
if len(sys.argv) != 2:
    print("Usage: main.py file.xlsx")
    print(sys.argv)
    exit(1)

[script_name, input_excel_file] = sys.argv
shifts_data = defaultdict(lambda: defaultdict(dict))
artist_names = set()

# Load workbook
input_workbook = load_workbook(filename=input_excel_file)

# Create folder if not exists
os.makedirs("output", exist_ok=True)

# Process sheets
for sheet in input_workbook:
    if not sheet.title.startswith("W"):
        continue

    print(sheet.title)

    season_info = sheet["A1"].value
    season_year = int(season_info.split(" ")[1].split("/")[0])

    calendar_week_info = sheet["A4"].value
    calendar_week_value = int(calendar_week_info.split(" ")[1])
    season_week = int(sheet["A6"].value)
    if season_week > calendar_week_value:
        season_year += 1

    days_row = sheet[3]
    hours_row = sheet[4]
    type_row = sheet[5]
    show_row = sheet[6]
    points_row = sheet[7]

    days_cell_range = sheet['D3':'BG3'][0]
    previous_date = None
    artist_days = []
    for cell in days_cell_range:
        day_value = cell.value
        if day_value is not None:
            day_parts = day_value.split(" ")
            current_date = datetime.strptime(f"{day_parts[1]}.{season_year}", DATE_FORMAT)
            previous_date = current_date
        artist_days.append(previous_date)

    for row in sheet.iter_rows(min_row=9, min_col=1, max_col=60):
        artist_name = row[0].value
        for cell in row:
            if cell.value not in ['D', 'DK', 'E', 'EK']: #D = Dienst, DK = Dienst Krank, #E = Ersatz, EK = Ersatz Krank. F = Fiktive Dienst (not for planning)
                continue
            shift_hours = str(hours_row[cell.column - 1].value)
            shift_day = artist_days[cell.column - 4]
            shifts_data[shift_day.isoformat()][artist_name][shift_hours] = [
                type_row[cell.column - 1].value or 'VST',
                show_row[cell.column - 1].value or '',
                points_row[cell.column - 1].value,
                cell.column
                ]
            artist_names.add(artist_name)

# Process artists and create calendar files
for artist_name in artist_names:
    c = Calendar()
    for shift_day in sorted(shifts_data):
        shifts = shifts_data[shift_day]
        artist_shifts = shifts[artist_name]
        sorted_shifts = sorted(artist_shifts)

        for shift_hours in sorted_shifts:
            [shift_type, show_info, points, column] = artist_shifts[shift_hours]
            date1 = shift_day.split('T')[0]
            hour1 = parse_hour(shift_hours)
            duration = dienst_duration(show_info, points)
            if duration is None:
                continue
            shift_start = datetime.strptime(f"{date1} {hour1}", "%Y-%m-%d %H.%M").replace(tzinfo=TIMEZONE)
            shift_end = shift_start + duration

            other_artists = [a for a in artist_names if a != artist_name and shift_hours in shifts[a]]

            e = Event()
            e.uid = str(uuid.uuid5(UUID_NAMESPACE, f"{artist_name}-{shift_day}-{column}"))
            e.name = f"{shift_type}: {show_info}"
            e.begin = shift_start
            e.end = shift_end
            e.description = f"Dienstwert: {points}\nWith: {', '.join(other_artists)}"
            c.events.add(e)

    with open(f"output/{artist_name}.ics", "w") as f:
        f.writelines(c.serialize_iter())

min_shift_date = datetime.fromisoformat(min(shifts_data.keys()))
max_shift_date = datetime.fromisoformat(max(shifts_data.keys()))

# Workbook initialization
output_workbook = Workbook()
first_sheet = output_workbook.active

# Process artists and create sheets
for artist_name in sorted(artist_names, key=lambda k: -len(k)):
    current_worksheet = output_workbook.create_sheet(title=' '.join(artist_name.split()[:2]))
    for shift_day in iterate_days(min_shift_date, max_shift_date + timedelta(days=90)):
        artist_shifts = shifts_data[shift_day.isoformat()]
        shift_day_date = shift_day.date()
        if artist_name not in artist_shifts:
            add_row(current_worksheet, [shift_day_date], COLOR1)
            continue
        sorted_shifts = sorted(artist_shifts[artist_name])

        row_hours = []
        row_types = []
        row_info = []
        row_points = 0
        for shift_hours in sorted_shifts:
            [shift_type, show_info, points, column] = artist_shifts[artist_name][shift_hours]
            row_hours.append(shift_hours.replace('.', ':'))
            row_types.append(shift_type)
            row_info.append(show_info)
            if points:
                row_points += points

        add_row(current_worksheet, [shift_day_date, merge_strings(row_hours), merge_strings(row_types), merge_strings(row_info), row_points], COLOR2)

    for row_data in current_worksheet.iter_rows():
        if row_data[0].value is None:
            continue
        shift_day_is_monday = row_data[0].value.weekday() == 0
        for cell in row_data:
            cell.border = THICK_BORDER if shift_day_is_monday else THIN_BORDER
        if contains_substring(row_data[2].value or '', ['VST', 'GP', 'WA', 'OHP', 'Pr-A', 'Pr-B']):
            row_data[2].font = BOLD_FONT
            row_data[3].font = BOLD_FONT

    current_worksheet.column_dimensions["A"].width = 10
    current_worksheet.column_dimensions["B"].width = 7
    current_worksheet.column_dimensions["C"].width = 5
    current_worksheet.column_dimensions["D"].width = 20

# Save workbook
output_workbook.remove(first_sheet)
output_workbook.save("output/dienstplan.xlsx")
